#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
建立 Renamer 工具的 Excel 模板檔案
包含「指令區」和「檔名變更區」兩個工作表
"""

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    print("使用 openpyxl 建立 Excel 檔案...")
    use_openpyxl = True
except ImportError:
    print("openpyxl 未安裝，使用基本的 pandas 功能...")
    import pandas as pd
    use_openpyxl = False

def create_excel_template():
    """建立完整的 Excel 模板檔案"""
    
    if use_openpyxl:
        create_with_openpyxl()
    else:
        create_with_pandas()

def create_with_openpyxl():
    """使用 openpyxl 建立功能完整的 Excel 檔案"""
    
    # 建立新的工作簿
    wb = Workbook()
    
    # 移除預設工作表
    wb.remove(wb.active)
    
    # 建立指令區工作表
    command_sheet = wb.create_sheet("指令區")
    setup_command_sheet(command_sheet)
    
    # 建立檔名變更區工作表
    filelist_sheet = wb.create_sheet("檔名變更區")
    setup_filelist_sheet(filelist_sheet)
    
    # 儲存檔案
    wb.save('renamer-template.xlsx')
    print("✅ Excel 模板檔案已建立：renamer-template.xlsx")

def setup_command_sheet(ws):
    """設定指令區工作表"""
    
    # 定義顏色
    title_fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")
    label_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # 定義字型
    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    
    # 定義對齊方式
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # 定義邊框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 設定標題
    ws.merge_cells('A1:B1')
    ws['A1'] = "檔案重新命名指令區"
    ws['A1'].fill = title_fill
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws['A1'].border = thin_border
    
    # 設定標籤和輸入欄位
    labels = [
        ("來源資料夾：", "請輸入資料夾ID或URL"),
        ("目標資料夾（複製模式）：", ""),
        ("重新命名模式：", "新增文字"),
        ("參數設定：", "前綴：IMG_"),
        ("操作類型：", "原位置更名")
    ]
    
    for i, (label, value) in enumerate(labels, start=2):
        # 標籤欄
        ws[f'A{i}'] = label
        ws[f'A{i}'].fill = label_fill
        ws[f'A{i}'].font = label_font
        ws[f'A{i}'].alignment = right_align
        ws[f'A{i}'].border = thin_border
        
        # 輸入欄
        ws[f'B{i}'] = value
        ws[f'B{i}'].font = normal_font
        ws[f'B{i}'].alignment = left_align
        ws[f'B{i}'].border = thin_border
    
    # 設定資料驗證
    # 重新命名模式下拉選單
    rename_mode_validation = DataValidation(
        type="list",
        formula1='"新增文字,取代文字,大小寫轉換,新增序號,格式化日期"',
        allow_blank=False
    )
    ws.add_data_validation(rename_mode_validation)
    rename_mode_validation.add(ws['B4'])
    
    # 操作類型下拉選單
    operation_validation = DataValidation(
        type="list",
        formula1='"原位置更名,複製後更名"',
        allow_blank=False
    )
    ws.add_data_validation(operation_validation)
    operation_validation.add(ws['B6'])
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40

def setup_filelist_sheet(ws):
    """設定檔名變更區工作表"""
    
    # 定義顏色
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    
    # 定義字型
    header_font = Font(bold=True, size=12, color="FFFFFF")
    normal_font = Font(size=10)
    
    # 定義對齊方式
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    # 定義邊框
    white_border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 設定標題列
    headers = [
        "原檔名", "原檔案路徑", "變更後檔名", "變更後檔案路徑", 
        "檔案類型", "檔案大小 (bytes)", "最後修改時間"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = white_border
    
    # 設定欄寬
    column_widths = [20, 25, 20, 25, 15, 12, 20]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # 凍結第一列
    ws.freeze_panes = 'A2'
    
    # 新增一些範例資料
    sample_data = [
        ["document.pdf", "測試資料夾/document.pdf", "IMG_document.pdf", "測試資料夾/IMG_document.pdf", "application/pdf", "1024", "2024/03/15 10:30:00"],
        ["image.jpg", "測試資料夾/image.jpg", "IMG_image.jpg", "測試資料夾/IMG_image.jpg", "image/jpeg", "2048", "2024/03/16 14:20:00"],
    ]
    
    for row, data in enumerate(sample_data, start=2):
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
            
            # 設定對齊方式
            if col in [5, 6]:  # 檔案大小和日期欄位
                cell.alignment = right_align if col == 6 else center_align
            else:
                cell.alignment = left_align

def create_with_pandas():
    """使用 pandas 建立基本的 Excel 檔案"""
    
    # 建立指令區資料
    command_data = {
        '標籤': ['檔案重新命名指令區', '來源資料夾：', '目標資料夾（複製模式）：', '重新命名模式：', '參數設定：', '操作類型：'],
        '值': ['', '請輸入資料夾ID或URL', '', '新增文字', '前綴：IMG_', '原位置更名']
    }
    
    # 建立檔名變更區資料
    filelist_data = {
        '原檔名': ['document.pdf', 'image.jpg'],
        '原檔案路徑': ['測試資料夾/document.pdf', '測試資料夾/image.jpg'],
        '變更後檔名': ['IMG_document.pdf', 'IMG_image.jpg'],
        '變更後檔案路徑': ['測試資料夾/IMG_document.pdf', '測試資料夾/IMG_image.jpg'],
        '檔案類型': ['application/pdf', 'image/jpeg'],
        '檔案大小 (bytes)': [1024, 2048],
        '最後修改時間': ['2024/03/15 10:30:00', '2024/03/16 14:20:00']
    }
    
    # 建立 DataFrame
    command_df = pd.DataFrame(command_data)
    filelist_df = pd.DataFrame(filelist_data)
    
    # 寫入 Excel 檔案
    with pd.ExcelWriter('renamer-template.xlsx', engine='openpyxl') as writer:
        command_df.to_excel(writer, sheet_name='指令區', index=False)
        filelist_df.to_excel(writer, sheet_name='檔名變更區', index=False)
    
    print("✅ 基本 Excel 模板檔案已建立：renamer-template.xlsx")

if __name__ == "__main__":
    create_excel_template()